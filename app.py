# 在文件的最开头添加这些代码
import os
os.environ['GRADIO_ANALYTICS_ENABLED'] = 'False'
os.environ['GRADIO_OFFILINE'] = 'True'

import sys

# 添加 py311 目录到 Python 路径
py311_path = os.path.join(os.path.dirname(__file__), 'py311')
sys.path.append(py311_path)

import gradio as gr
gr.close_all()

import ssl
# 仅在 SSL 证书验证失败时才跳过验证，受控环境下使用
try:
    ssl.create_default_context()
except Exception:
    ssl._create_default_https_context = ssl._create_unverified_context

from ultralytics import YOLO
import cv2
import numpy as np
import torch
import tempfile
import subprocess
import shutil
import json
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)

# COCO 数据集 17 个关键点标准名称（姿态估计）
COCO_KEYPOINT_NAMES = [
    "nose",              # 0  鼻
    "left_eye",          # 1  左眼
    "right_eye",         # 2  右眼
    "left_ear",          # 3  左耳
    "right_ear",         # 4  右耳
    "left_shoulder",     # 5  左肩
    "right_shoulder",    # 6  右肩
    "left_elbow",        # 7  左肘
    "right_elbow",       # 8  右肘
    "left_wrist",        # 9  左腕
    "right_wrist",       # 10 右腕
    "left_hip",          # 11 左髋
    "right_hip",         # 12 右髋
    "left_knee",         # 13 左膝
    "right_knee",        # 14 右膝
    "left_ankle",        # 15 左踝
    "right_ankle",       # 16 右踝
]

# COCO 关键点中英对照（用于 Excel 表头）
COCO_KEYPOINT_NAMES_CN = {
    "nose": "鼻子",
    "left_eye": "左眼",
    "right_eye": "右眼",
    "left_ear": "左耳",
    "right_ear": "右耳",
    "left_shoulder": "左肩",
    "right_shoulder": "右肩",
    "left_elbow": "左肘",
    "right_elbow": "右肘",
    "left_wrist": "左腕",
    "right_wrist": "右腕",
    "left_hip": "左髋",
    "right_hip": "右髋",
    "left_knee": "左膝",
    "right_knee": "右膝",
    "left_ankle": "左踝",
    "right_ankle": "右踝",
}

# 检查CUDA是否可用
cuda_available = torch.cuda.is_available()

# 初始化YOLO模型
models = {}

ONLINE_MODELS = {
    "检测": ["yolo11n.pt", "yolo11s.pt", "yolo11m.pt", "yolo11l.pt", "yolo11x.pt"],
    "分割": ["yolo11n-seg.pt", "yolo11s-seg.pt", "yolo11m-seg.pt", "yolo11l-seg.pt", "yolo11x-seg.pt"],
    "姿态": ["yolo11n-pose.pt", "yolo11s-pose.pt", "yolo11m-pose.pt", "yolo11l-pose.pt", "yolo11x-pose.pt"],
    "分类": ["yolo11n-cls.pt", "yolo11s-cls.pt", "yolo11m-cls.pt", "yolo11l-cls.pt", "yolo11x-cls.pt"],
    "OBB": ["yolo11n-obb.pt", "yolo11s-obb.pt", "yolo11m-obb.pt", "yolo11l-obb.pt", "yolo11x-obb.pt"]
}

def get_available_models(task):
    models_dir = "models"
    if not os.path.exists(models_dir):
        os.makedirs(models_dir)
    
    task_folders = {
        "检测": "detection",
        "分割": "segmentation",
        "姿态": "pose",
        "分类": "classification",
        "OBB": "obb"
    }
    
    task_folder = task_folders.get(task)
    if not task_folder:
        return []
    
    task_path = os.path.join(models_dir, task_folder)
    if not os.path.exists(task_path):
        os.makedirs(task_path)
    
    local_models = set(f for f in os.listdir(task_path) if f.endswith('.pt'))
    online_models = set(ONLINE_MODELS.get(task, []))
    
    all_models = local_models.union(online_models)
    return sorted(list(all_models))

def get_model(task, model_name, device):
    task_folders = {
        "检测": "detection",
        "分割": "segmentation",
        "姿态": "pose",
        "分类": "classification",
        "OBB": "obb"
    }
    task_folder = task_folders.get(task)
    if not task_folder:
        raise ValueError(f"不支持的任务类型: {task}")
    
    model_path = os.path.join("models", task_folder, model_name)
    if not os.path.exists(model_path):
        model_path = download_model(task, model_name)
    
    if (task, model_name) not in models:
        models[(task, model_name)] = YOLO(model_path)
    if device == "cuda" and cuda_available:
        models[(task, model_name)].to(device)
    return models[(task, model_name)]

def split_dataset(data_path, test_size=0.2, val_size=0.2):
    from sklearn.model_selection import train_test_split

    images_dir = os.path.join(data_path, 'images')
    labels_dir = os.path.join(data_path, 'labels')
    
    if not os.path.exists(images_dir):
        raise ValueError(f"图像目录不存在: {images_dir}")
    
    all_images = [f for f in os.listdir(images_dir) if f.endswith(('.jpg', '.png', '.jpeg'))]
    
    if len(all_images) == 0:
        raise ValueError(f"在 {images_dir} 中没有找到图像文件")
    
    # 首先分割出验证集和测试集
    train_images, val_test_images = train_test_split(all_images, test_size=(test_size + val_size), random_state=42)
    val_images, test_images = train_test_split(val_test_images, test_size=val_size/(test_size + val_size), random_state=42)
    
    # 创建新的目录结构
    for split in ['train', 'val', 'test']:
        os.makedirs(os.path.join(data_path, 'images', split), exist_ok=True)
        os.makedirs(os.path.join(data_path, 'labels', split), exist_ok=True)
    
    # 复制文件
    for split, image_list in [('train', train_images), ('val', val_images), ('test', test_images)]:
        for img in image_list:
            shutil.copy2(os.path.join(images_dir, img), os.path.join(data_path, 'images', split, img))
            label = img.rsplit('.', 1)[0] + '.txt'
            if os.path.exists(os.path.join(labels_dir, label)):
                shutil.copy2(os.path.join(labels_dir, label), os.path.join(data_path, 'labels', split, label))

def create_dataset_yaml(data_path, include_test=False):
    import yaml
    class_names = get_class_names(data_path)
    yaml_content = {
        'path': os.path.abspath(data_path),  # 使用绝对路径
        'train': 'images/train',
        'val': 'images/val',
        'nc': len(class_names),
        'names': class_names
    }
    
    if include_test:
        yaml_content['test'] = 'images/test'
    
    yaml_path = os.path.join(data_path, 'dataset.yaml')
    with open(yaml_path, 'w', encoding='utf-8') as file:
        yaml.dump(yaml_content, file, sort_keys=False, allow_unicode=True)
    
    return yaml_path

def get_class_names(data_path):
    classes_file = os.path.join(data_path, 'classes', 'classes.txt')
    if os.path.exists(classes_file):
        encodings = ['utf-8', 'gbk', 'gb2312', 'iso-8859-1']
        for encoding in encodings:
            try:
                with open(classes_file, 'r', encoding=encoding) as f:
                    class_names = [line.strip() for line in f.readlines() if line.strip()]
                return class_names
            except UnicodeDecodeError:
                continue
        logging.warning(f"无法以支持的编码读取类别文件: {classes_file}")
        return []
    else:
        logging.warning(f"未找到类别文件: {classes_file}")
        return []

def get_ultralytics_settings():
    settings_path = os.path.join(os.getcwd(), 'ultralytics_settings.json')
    if os.path.exists(settings_path):
        with open(settings_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'datasets_dir': './datasets'}

def check_dataset(data_path):
    images_dir = os.path.join(data_path, 'images')
    labels_dir = os.path.join(data_path, 'labels')
    classes_file = os.path.join(data_path, 'classes', 'classes.txt')

    # 检查目录结构
    if not os.path.exists(images_dir) or not os.path.exists(labels_dir):
        logging.error(f"数据集结构不正确。请确保 {data_path} 目录下有 'images' 和 'labels' 子目录。")
        return False

    # 检查类别文件
    if not os.path.exists(classes_file):
        logging.error(f"未找到类别文件: {classes_file}")
        return False

    # 检查图像和标签文件的对应关系
    image_files = set(f for f in os.listdir(images_dir) if f.endswith(('.jpg', '.png', '.jpeg')))
    label_files = set(f for f in os.listdir(labels_dir) if f.endswith('.txt'))
    
    if len(image_files) == 0:
        logging.error(f"在 {images_dir} 中没有找到图像文件")
        return False

    for img_file in image_files:
        label_file = os.path.splitext(img_file)[0] + '.txt'
        if label_file not in label_files:
            logging.warning(f"图像文件 {img_file} 没有对应的标签文件")

    # 检查标签文件格式
    class_names = get_class_names(data_path)
    for label_file in label_files:
        label_path = os.path.join(labels_dir, label_file)
        # 尝试多种编码读取标签文件
        label_content = None
        for encoding in ['utf-8', 'gbk', 'gb2312', 'iso-8859-1']:
            try:
                with open(label_path, 'r', encoding=encoding) as f:
                    label_content = f.readlines()
                break
            except UnicodeDecodeError:
                continue
        if label_content is None:
            logging.error(f"无法以支持的编码读取标签文件: {label_file}")
            return False

        for line in label_content:
            parts = line.strip().split()
            if len(parts) != 5:
                logging.error(f"标签文件 {label_file} 格式不正确")
                return False
            try:
                class_id = int(parts[0])
            except ValueError:
                logging.error(f"标签文件 {label_file} 中的类别ID不是整数: {parts[0]}")
                return False
            if class_id >= len(class_names):
                logging.error(f"标签文件 {label_file} 中的类别ID {class_id} 超出范围")
                return False
            try:
                x, y, w, h = map(float, parts[1:])
            except ValueError:
                logging.error(f"标签文件 {label_file} 中的坐标格式不正确")
                return False
            if not all(0 <= v <= 1 for v in (x, y, w, h)):
                logging.error(f"标签文件 {label_file} 中的坐标不在0-1范围内")
                return False

    logging.info("数据集检查通过")
    return True

def train_model(data_path, epochs, task, model_name, device, include_test, test_size, val_size, 
                batch_size, imgsz, patience, optimizer, lr0, lrf, momentum, weight_decay, 
                warmup_epochs, warmup_momentum, warmup_bias_lr, box, cls, dfl, pose, kobj, 
                label_smoothing, nbs, overlap_mask, mask_ratio, dropout,
                # 新增的参数
                augment, mixup, copy_paste, mosaic, amp):
    try:
        if not check_dataset(data_path):
            return "数据集检查失败，请查看日志以获取详细信息。"

        settings = get_ultralytics_settings()
        datasets_dir = settings.get('datasets_dir', '.')
        
        if not os.path.isabs(data_path):
            data_path = os.path.abspath(data_path)
        
        if not os.path.exists(data_path):
            return f"错误：数据集路径不存在 - {data_path}"
        
        if not (os.path.exists(os.path.join(data_path, 'images')) and 
                os.path.exists(os.path.join(data_path, 'labels'))):
            return f"错误：数据集结构不正确。请确保 {data_path} 目录下有 'images' 和 'labels' 子目录。"
        
        split_dataset(data_path, test_size=float(test_size), val_size=float(val_size))
        
        yaml_path = create_dataset_yaml(data_path, include_test=include_test)
        
        model = get_model(task, model_name, device)
        
        torch.cuda.empty_cache()  # 在训练开始前清理GPU缓存
        
        results = model.train(
            data=yaml_path, 
            epochs=int(epochs), 
            device=device,
            batch=int(batch_size),
            imgsz=int(imgsz),
            patience=int(patience),
            optimizer=optimizer,
            lr0=float(lr0),
            lrf=float(lrf),
            momentum=float(momentum),
            weight_decay=float(weight_decay),
            warmup_epochs=float(warmup_epochs),
            warmup_momentum=float(warmup_momentum),
            warmup_bias_lr=float(warmup_bias_lr),
            box=float(box),
            cls=float(cls),
            dfl=float(dfl),
            pose=float(pose),
            kobj=float(kobj),
            label_smoothing=float(label_smoothing),
            nbs=int(nbs),
            overlap_mask=overlap_mask,
            mask_ratio=int(mask_ratio),
            dropout=float(dropout),
            cache=True,
            amp=amp,
            augment=augment,
            mixup=float(mixup),
            copy_paste=float(copy_paste),
            mosaic=float(mosaic),
        )
        
        # 提取关键信息
        metrics = results.results_dict
        best_fitness = metrics.get('fitness', 'N/A')
        mAP50 = metrics.get('metrics/mAP50(B)', 'N/A')
        mAP50_95 = metrics.get('metrics/mAP50-95(B)', 'N/A')
        
        # 获取模型保存路径
        save_dir = results.save_dir
        best_model_path = os.path.join(save_dir, 'weights', 'best.pt')
        
        # 构建简化的结果字符串
        result_str = f"训练完成。\n"
        result_str += f"最佳fitness: {best_fitness}\n"
        result_str += f"mAP50: {mAP50}\n"
        result_str += f"mAP50-95: {mAP50_95}\n"
        result_str += f"最佳模型保存路径: {best_model_path}"
        
        return result_str
    except Exception as e:
        logging.error(f"训练过程中出现错误: {str(e)}", exc_info=True)
        return f"训练过程中出现错误: {str(e)}"

def validate_model(data, task, model_name, device, batch_size, imgsz, conf, iou, max_det, half, plots):
    model = get_model(task, model_name, device)
    results = model.val(data=data, device=device, batch=batch_size, imgsz=imgsz, conf=conf, iou=iou, max_det=max_det, half=half, plots=plots)
    return f"验证完成。结果: {results}"

def translate_label(label):
    # 扩展翻译字典,包含更多常见物体
    translations = {
        'person': '人',
        'bicycle': '自行车',
        'car': '汽车',
        'motorcycle': '摩托车',
        'airplane': '飞机',
        'bus': '公交车',
        'train': '火车',
        'truck': '卡车',
        'boat': '船',
        'traffic light': '交通灯',
        'fire hydrant': '消防栓',
        'stop sign': '停止标志',
        'parking meter': '停车计时器',
        'bench': '长凳',
        'bird': '鸟',
        'cat': '猫',
        'dog': '狗',
        'horse': '马',
        'sheep': '羊',
        'cow': '牛',
        'elephant': '大象',
        'bear': '熊',
        'zebra': '斑马',
        'giraffe': '长颈鹿',
        'backpack': '背包',
        'umbrella': '雨伞',
        'handbag': '手提包',
        'tie': '领带',
        'suitcase': '行李箱',
        'frisbee': '飞盘',
        'skis': '滑雪板',
        'snowboard': '单板滑雪',
        'sports ball': '运动球',
        'kite': '风筝',
        'baseball bat': '棒球棒',
        'baseball glove': '棒球手套',
        'skateboard': '滑板',
        'surfboard': '冲浪板',
        'tennis racket': '网球拍',
        'bottle': '瓶子',
        'wine glass': '酒杯',
        'cup': '杯子',
        'fork': '叉子',
        'knife': '刀',
        'spoon': '勺子',
        'bowl': '碗',
        'banana': '香蕉',
        'apple': '苹果',
        'sandwich': '三明治',
        'orange': '橙子',
        'broccoli': '西兰花',
        'carrot': '胡萝卜',
        'hot dog': '热狗',
        'pizza': '披萨',
        'donut': '甜甜圈',
        'cake': '蛋糕',
        'chair': '椅子',
        'couch': '沙发',
        'potted plant': '盆栽',
        'bed': '床',
        'dining table': '餐桌',
        'toilet': '马桶',
        'tv': '电视',
        'laptop': '笔记本电脑',
        'mouse': '鼠标',
        'remote': '遥控器',
        'keyboard': '键盘',
        'cell phone': '手机',
        'microwave': '微波炉',
        'oven': '烤箱',
        'toaster': '烤面包机',
        'sink': '水槽',
        'refrigerator': '冰箱',
        'book': '书',
        'clock': '时钟',
        'vase': '花瓶',
        'scissors': '剪刀',
        'teddy bear': '泰迪熊',
        'hair drier': '吹风机',
        'toothbrush': '牙刷',
    }
    return translations.get(label.lower(), label)  # 如果没有翻译,返回原标签

def _parse_classes_param(classes):
    """解析类别过滤参数，返回列表或None"""
    if classes:
        try:
            return [int(c.strip()) for c in classes.split(',') if c.strip()]
        except ValueError:
            return None  # 格式错误时返回None，由调用方处理
    return None

def predict_image(image, task, model_name, device, conf, iou, max_det, classes, agnostic_nms, augment):
    model = get_model(task, model_name, device)
    
    # Gradio 以 RGB 格式提供图像，YOLO 原生接受 BGR 格式
    # 因此需要将 RGB 转为 BGR 供 YOLO 处理
    if len(image.shape) == 3 and image.shape[-1] == 3:
        image_bgr = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
    else:
        image_bgr = image  # 灰度图像保持不变
    
    # 处理 classes 参数
    parsed_classes = _parse_classes_param(classes)
    if classes and parsed_classes is None:
        return None, {"error": "类别过滤参数格式不正确。请使用逗号分隔的整数。"}, None
    
    try:
        results = model(image_bgr, device=device, conf=conf, iou=iou, max_det=max_det, classes=parsed_classes, agnostic_nms=agnostic_nms, augment=augment)
    
        plot_result = None
        json_result = None
        excel_path = None  # 姿态任务时生成的 Excel 数据文件
        
        if task in ["检测", "分割", "姿态", "OBB"]:
            # 修改标签为中文
            for r in results:
                r.names = {k: translate_label(v) for k, v in r.names.items()}
                
                # 对于检测任务,我们需要更新每个检测框的标签
                if task == "检测":
                    for box in r.boxes:
                        if box.cls.numel() > 0:
                            class_id = int(box.cls[0])
                            box.label = translate_label(r.names[class_id])
            
            plot_result = results[0].plot()
            
            # 姿态任务：同时导出关键点数据到 Excel
            if task == "姿态":
                excel_path = _export_pose_data_from_results(results, "uploaded_image")
        elif task == "分类":
            names = results[0].names
            probs = results[0].probs.data.tolist()
            json_result = {translate_label(names[i]): probs[i] for i in range(len(names))}
        
        if plot_result is not None:
            if isinstance(plot_result, torch.Tensor):
                plot_result = plot_result.cpu().numpy()
            
            # results.plot() 输出 BGR 格式，转为 RGB 供 Gradio 显示
            if len(plot_result.shape) == 3 and plot_result.shape[-1] == 3:
                plot_result = cv2.cvtColor(plot_result, cv2.COLOR_BGR2RGB)
        
        return plot_result, json_result, excel_path
    except Exception as e:
        return None, {"error": f"预测过程中出现错误: {str(e)}"}, None


def _export_pose_data_from_results(results, image_name):
    """从已有推理结果中提取姿态数据并导出 Excel，返回文件路径或 None"""
    try:
        import openpyxl
    except ImportError:
        logging.warning("缺少 openpyxl 库，无法导出姿态数据")
        return None

    rows = _extract_pose_rows(results, image_name)
    if not rows:
        return None

    output_dir = os.path.join(os.getcwd(), "exports")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"pose_export_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    _write_pose_excel(rows, excel_path)
    return excel_path


def _export_pose_data_from_rows(all_rows):
    """从已收集的姿态数据行导出 Excel，返回文件路径或 None"""
    try:
        import openpyxl
    except ImportError:
        logging.warning("缺少 openpyxl 库，无法导出姿态数据")
        return None

    if not all_rows:
        return None

    output_dir = os.path.join(os.getcwd(), "exports")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"pose_export_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    _write_pose_excel(all_rows, excel_path)
    return excel_path

def export_model(format, task, model_name, device, imgsz, batch, half, int8, dynamic, simplify, opset):
    model = get_model(task, model_name, device)
    success = model.export(format=format, imgsz=imgsz, batch=batch, half=half, int8=int8, dynamic=dynamic, simplify=simplify, opset=opset)
    return f"模型导出{'成功' if success else '失败'}"

def export_pose_data(source, source_type, model_name, device, conf, iou, max_det, output_dir):
    """
    从姿态估计模型的推理结果中导出原始数据到 Excel。
    
    导出内容：
    - 边界框: [x1, y1, x2, y2] 像素坐标
    - 检测置信度
    - 所有关键点: 像素坐标 x, y + 置信度 + 关键点名称
    
    参数:
        source: 图像路径(str)或numpy数组
        source_type: "image" 或 "video"
        model_name: 模型名称
        device: 设备
        conf: 置信度阈值
        iou: IOU阈值
        max_det: 最大检测数
        output_dir: 导出目录
    
    返回:
        (excel_path, summary_str) 或 (None, error_str)
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    try:
        model = get_model("姿态", model_name, device)
        
        # 确保输出目录存在
        if not output_dir:
            output_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(output_dir, exist_ok=True)

        all_rows = []  # 收集所有数据行
        image_names = []  # 对应的图像/帧标识

        if source_type == "image":
            # 单张图像推理
            if isinstance(source, np.ndarray):
                # Gradio 传来的 numpy 数组（RGB格式）
                if len(source.shape) == 3 and source.shape[-1] == 3:
                    image_bgr = cv2.cvtColor(source, cv2.COLOR_RGB2BGR)
                else:
                    image_bgr = source
                img_name = "uploaded_image"
                results = model(image_bgr, device=device, conf=conf, iou=iou, max_det=max_det)
            elif isinstance(source, str):
                img_name = os.path.basename(source)
                results = model(source, device=device, conf=conf, iou=iou, max_det=max_det)
            else:
                return None, f"不支持的输入类型: {type(source)}"

            rows = _extract_pose_rows(results, img_name)
            all_rows.extend(rows)
            image_names.append(img_name)

        elif source_type == "video":
            # 视频逐帧推理
            video_path = source
            if isinstance(video_path, dict) and 'name' in video_path:
                video_path = video_path['name']
            elif not isinstance(video_path, str):
                return None, f"无效的视频路径类型: {type(video_path)}"

            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                return None, "无法打开视频文件"

            frame_idx = 0
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            video_name = os.path.basename(video_path)

            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    break
                
                frame_idx += 1
                img_name = f"{video_name}_frame_{frame_idx:06d}"
                results = model(frame, device=device, conf=conf, iou=iou, max_det=max_det)
                rows = _extract_pose_rows(results, img_name)
                all_rows.extend(rows)

                if frame_idx % 50 == 0:
                    logging.info(f"视频帧处理进度: {frame_idx}/{total_frames}")

            cap.release()
            image_names.append(f"{video_name} ({frame_idx} 帧)")

        else:
            return None, f"不支持的源类型: {source_type}"

        if not all_rows:
            return None, "未检测到任何姿态目标，无数据可导出。请尝试降低置信度阈值。"

        # 生成 Excel 文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"pose_export_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)

        _write_pose_excel(all_rows, excel_path)

        # 汇总信息
        num_images = len(set(r[0] for r in all_rows))  # 第0列是图像名
        num_persons = len(all_rows)
        summary = (
            f"✅ 数据导出成功！\n\n"
            f"📁 文件路径: {excel_path}\n\n"
            f"📊 统计信息:\n"
            f"  • 图像/帧数: {num_images}\n"
            f"  • 检测到的人体数量: {num_persons}\n"
            f"  • 每人关键点数: 17 (COCO标准)\n"
            f"  • 关键点名称: {', '.join(COCO_KEYPOINT_NAMES)}"
        )
        return excel_path, summary

    except ImportError:
        return None, "缺少 openpyxl 库，请运行: pip install openpyxl"
    except Exception as e:
        logging.error(f"数据导出出错: {str(e)}", exc_info=True)
        return None, f"数据导出出错: {str(e)}"


def _extract_pose_rows(results, image_name):
    """
    从 YOLO 姿态推理结果中提取一行一行的原始数据。
    
    每行格式: [图像名, 人体编号, x1, y1, x2, y2, 检测置信度, 
               kp0_name, kp0_x, kp0_y, kp0_conf, kp1_name, kp1_x, kp1_y, kp1_conf, ...]
    """
    rows = []
    for r in results:
        if r.boxes is None or len(r.boxes) == 0:
            continue

        boxes = r.boxes
        keypoints = r.keypoints

        for person_idx in range(len(boxes)):
            # 边界框 [x1, y1, x2, y2]
            box = boxes[person_idx]
            xyxy = box.xyxy[0].cpu().numpy()
            x1, y1, x2, y2 = float(xyxy[0]), float(xyxy[1]), float(xyxy[2]), float(xyxy[3])

            # 检测置信度
            det_conf = float(box.conf[0].cpu()) if box.conf.numel() > 0 else 0.0

            row = [image_name, person_idx + 1, x1, y1, x2, y2, det_conf]

            # 关键点
            if keypoints is not None and len(keypoints) > person_idx:
                kpts = keypoints[person_idx]
                kpts_xy = kpts.xy[0].cpu().numpy()   # shape: (17, 2)
                kpts_conf = kpts.conf[0].cpu().numpy() if kpts.conf is not None else np.ones(17)  # shape: (17,)

                for kp_idx in range(min(len(kpts_xy), len(COCO_KEYPOINT_NAMES))):
                    kp_name = COCO_KEYPOINT_NAMES[kp_idx]
                    kp_x = float(kpts_xy[kp_idx][0])
                    kp_y = float(kpts_xy[kp_idx][1])
                    kp_c = float(kpts_conf[kp_idx]) if kp_idx < len(kpts_conf) else 0.0
                    row.extend([kp_name, kp_x, kp_y, kp_c])
            else:
                # 没有关键点数据，填充占位
                for kp_name in COCO_KEYPOINT_NAMES:
                    row.extend([kp_name, 0.0, 0.0, 0.0])

            rows.append(row)

    return rows


def _write_pose_excel(rows, excel_path):
    """
    将姿态数据行写入 Excel 文件，带格式化表头。
    每行数据对应一个检测到的人体，关键点信息以列展开。
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "姿态数据"

    # 构建表头
    base_headers = ["图像名称", "人体编号", "x1", "y1", "x2", "y2", "检测置信度"]
    kp_headers = []
    for kp_name in COCO_KEYPOINT_NAMES:
        cn_name = COCO_KEYPOINT_NAMES_CN.get(kp_name, kp_name)
        kp_headers.extend([
            f"{kp_name}\n({cn_name})_x",
            f"{kp_name}\n({cn_name})_y", 
            f"{kp_name}\n({cn_name})_conf",
        ])
    headers = base_headers + kp_headers
    ws.append(headers)

    # 样式定义
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 应用表头样式
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据行
    data_font = Font(name="Consolas", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    for row_data in rows:
        # 重新排列: 去掉每4个中的kp_name（写入时名称已在表头），只保留数值
        base_data = row_data[:7]  # 图像名, 编号, x1, y1, x2, y2, 置信度
        kp_data = row_data[7:]    # 每4个: [name, x, y, conf]

        # 只提取 x, y, conf（name已在表头）
        kp_values = []
        for i in range(0, len(kp_data), 4):
            kp_values.extend([kp_data[i + 1], kp_data[i + 2], kp_data[i + 3]])

        full_row = base_data + kp_values
        ws.append(full_row)

        # 应用数据样式
        for col_idx in range(1, len(full_row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            # 数值列保留4位小数
            if col_idx >= 3 and isinstance(cell.value, float):
                cell.number_format = '0.0000'

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_length = max(max_length, cell_len)
        # 基础宽度，关键点列稍窄
        adjusted_width = min(max_length + 4, 22)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 12)

    # 冻结首行表头
    ws.freeze_panes = "A2"

    # 添加第二个 Sheet：关键点说明
    ws2 = wb.create_sheet("关键点说明")
    ws2.append(["关键点编号", "英文名称", "中文名称"])
    ref_font = Font(name="微软雅黑", size=10)
    for idx, name in enumerate(COCO_KEYPOINT_NAMES):
        cn = COCO_KEYPOINT_NAMES_CN.get(name, name)
        ws2.append([idx, name, cn])
        for col in range(1, 4):
            ws2.cell(row=idx + 2, column=col).font = ref_font
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 12

    wb.save(excel_path)
    logging.info(f"姿态数据已导出至: {excel_path}")

def process_video(video_path, task, model_name, device, conf, iou, max_det, classes, agnostic_nms, augment):
    try:
        logging.info(f"开始处理视频: {video_path}")

        if isinstance(video_path, dict) and 'name' in video_path:
            video_path = video_path['name']
        elif not isinstance(video_path, str):
            raise ValueError(f"无效的视频路径类型: {type(video_path)}")

        model = get_model(task, model_name, device)
        cap = cv2.VideoCapture(video_path)
        
        if not cap.isOpened():
            raise ValueError("无法打开视频文件")
        
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = round(cap.get(cv2.CAP_PROP_FPS))  # 使用 round 避免精度丢失
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.mp4') as temp_file:
            output_path = temp_file.name
        
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
        
        # classes 参数此时应为列表（由 on_video_button_click 解析），或 None
        if isinstance(classes, str):
            parsed_classes = _parse_classes_param(classes)
        else:
            parsed_classes = classes  # 已经是列表或None

        # 姿态任务时，收集所有帧的推理数据
        all_pose_rows = [] if task == "姿态" else None
        video_name = os.path.basename(video_path)

        frame_count = 0
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
            
            # cv2 读取的帧为 BGR 格式，YOLO 原生接受 BGR，直接送入模型
            results = model(frame, device=device, conf=conf, iou=iou, max_det=max_det, classes=parsed_classes, agnostic_nms=agnostic_nms, augment=augment)
            
            if task in ["检测", "分割", "姿态", "OBB"]:
                # 修改标签为中文
                for r in results:
                    r.names = {k: translate_label(v) for k, v in r.names.items()}
                    
                    # 对于检测任务,我们需要更新每个检测框的标签
                    if task == "检测":
                        for box in r.boxes:
                            if box.cls.numel() > 0:
                                class_id = int(box.cls[0])
                                box.label = translate_label(r.names[class_id])
                
                annotated_frame = results[0].plot()

                # 姿态任务：收集本帧的关键点数据
                if task == "姿态":
                    frame_count += 1
                    frame_name = f"{video_name}_frame_{frame_count:06d}"
                    rows = _extract_pose_rows(results, frame_name)
                    all_pose_rows.extend(rows)
            else:
                raise ValueError(f"不支持的任务类型: {task}")
            
            # results.plot() 输出 BGR 格式，可以直接写入 VideoWriter
            out.write(annotated_frame)
            if task != "姿态":
                frame_count += 1
        
        cap.release()
        out.release()
        logging.info("视频处理完成，开始进行 FFmpeg 编码")
        
        ffmpeg_path = r"FFmpeg/bin/ffmpeg.exe"
        if not os.path.exists(ffmpeg_path):
            raise ValueError(f"FFmpeg 可执行文件不存在: {ffmpeg_path}")

        final_output_path = output_path.replace('.mp4', '_final.mp4')
        
        ffmpeg_command = [
            ffmpeg_path,
            "-i", output_path,
            "-i", video_path,
            "-c:v", "libx264",
            "-preset", "slow",
            "-crf", "22",
            "-c:a", "aac",
            "-map", "0:v:0",
            "-map", "1:a:0?",
            final_output_path
        ]
        
        subprocess.run(ffmpeg_command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        os.remove(output_path)
        
        # 姿态任务：导出 Excel 数据文件
        excel_path = None
        if task == "姿态" and all_pose_rows:
            excel_path = _export_pose_data_from_rows(all_pose_rows)

        logging.info("视频处理完成")
        return final_output_path, excel_path
    except Exception as e:
        logging.error(f"处理视频时发生错误: {str(e)}")
        return None, None

def on_video_button_click(video_path, task, model_name, device, conf, iou, max_det, classes, agnostic_nms, augment):
    if video_path is None:
        return None, "请上传视频文件。", gr.update(visible=False, value=None)
    
    if isinstance(video_path, dict) and 'name' in video_path:
        video_path = video_path['name']
    elif not isinstance(video_path, str):
        return None, f"无效的视频路径类型: {type(video_path)}。请确保上传了正确的视频文件。", gr.update(visible=False, value=None)
    
    # 处理 classes 参数：在此处解析为列表，传给 process_video
    parsed_classes = _parse_classes_param(classes)
    if classes and parsed_classes is None:
        return None, "类别过滤参数格式不正确。请使用逗号分隔的整数。", gr.update(visible=False, value=None)
    
    result, excel_path = process_video(video_path, task, model_name, device, conf, iou, max_det, parsed_classes, agnostic_nms, augment)
    if result is None:
        return None, f"视频处理失败: {task}任务出错，请检查控制台输出以获取更多信息。", gr.update(visible=False, value=None)
    
    # 姿态任务且有数据时，显示 Excel 下载；否则隐藏
    if excel_path:
        return result, None, gr.update(visible=True, value=excel_path)
    else:
        return result, None, gr.update(visible=False, value=None)

def download_model(task, model_name):
    task_folders = {
        "检测": "detection",
        "分割": "segmentation",
        "姿态": "pose",
        "分类": "classification",
        "OBB": "obb"
    }
    task_folder = task_folders.get(task)
    if not task_folder:
        raise ValueError(f"不支持的任务类型: {task}")
    
    model_path = os.path.join("models", task_folder, model_name)
    if not os.path.exists(model_path):
        print(f"正在下载模型: {model_name}")
        
        # 确保目标文件夹存在
        os.makedirs(os.path.dirname(model_path), exist_ok=True)
        
        # 使用 YOLO 下载模型，指定保存路径避免在工作目录留下文件
        temp_model = YOLO(model_name)
        
        # 查找 YOLO 默认下载路径中的模型文件并复制到目标位置
        # YOLO 默认将模型下载到当前目录或 Ultralytics 设置的目录
        possible_sources = [
            model_name,  # 当前目录
            os.path.join(os.getcwd(), model_name),  # 绝对路径
        ]
        
        # 检查 Ultralytics 的默认下载路径
        try:
            from ultralytics.utils import SETTINGS
            ultralytics_dir = SETTINGS.get('weights_dir', '')
            if ultralytics_dir:
                possible_sources.append(os.path.join(ultralytics_dir, model_name))
        except Exception:
            pass
        
        source_found = False
        for src in possible_sources:
            if os.path.exists(src) and os.path.abspath(src) != os.path.abspath(model_path):
                shutil.copy2(src, model_path)
                # 清理下载源文件（仅当源文件在当前目录或临时位置时）
                if os.path.dirname(os.path.abspath(src)) == os.path.abspath(os.getcwd()):
                    try:
                        os.remove(src)
                    except OSError:
                        pass
                source_found = True
                break
        
        if not source_found:
            # 如果找不到下载的文件，尝试直接使用模型路径重新加载
            logging.warning(f"未能找到下载的模型源文件，尝试直接使用模型路径")
            # YOLO 可能已缓存模型，直接使用 model_path 检查
            if not os.path.exists(model_path):
                raise FileNotFoundError(f"模型下载后未找到文件: {model_name}")
        
        print(f"模型已保存到: {model_path}")
    
    return model_path

def update_model_choices(task):
    return gr.Dropdown(choices=get_available_models(task))

# Gradio界面
with gr.Blocks() as demo:
    gr.Markdown("# Ultralytics YOLO11 WebUI\n[官方GitHub项目地址](https://github.com/ultralytics/ultralytics)")
    
    device_options = ["cuda", "cpu"] if cuda_available else ["cpu"]
    device_input = gr.Radio(device_options, label="设备选择", value="cuda" if cuda_available else "cpu")
    
    with gr.Tab("训练"):
        train_task_input = gr.Dropdown(["检测", "分割", "姿态", "分类", "OBB"], label="任务类型")
        train_model_input = gr.Dropdown([], label="选择模型")
        data_input = gr.Textbox(label="数据集根目录路径", value="my_dataset")
        epochs_input = gr.Number(label="训练轮数", value=100)
        batch_size_input = gr.Number(label="批量大小", value=16)
        imgsz_input = gr.Number(label="图像尺寸", value=640)
        include_test = gr.Checkbox(label="包含测试集", value=True)
        test_size = gr.Slider(minimum=0.1, maximum=0.5, value=0.2, label="测试集比例")
        val_size = gr.Slider(minimum=0.1, maximum=0.5, value=0.2, label="验证集比例")
        
        with gr.Accordion("高级选项", open=False):
            patience_input = gr.Number(label="早停耐心值", value=50)
            optimizer_input = gr.Dropdown(["SGD", "Adam", "AdamW", "RMSProp"], label="优化器", value="SGD")
            lr0_input = gr.Number(label="初始学习率", value=0.01)
            lrf_input = gr.Number(label="最终学习率因子", value=0.01)
            momentum_input = gr.Number(label="动量", value=0.937)
            weight_decay_input = gr.Number(label="权重衰减", value=0.0005)
            warmup_epochs_input = gr.Number(label="预热轮数", value=3.0)
            warmup_momentum_input = gr.Number(label="预热动量", value=0.8)
            warmup_bias_lr_input = gr.Number(label="预热偏置学习率", value=0.1)
            box_input = gr.Number(label="框损失增益", value=7.5)
            cls_input = gr.Number(label="类别损失增益", value=0.5)
            dfl_input = gr.Number(label="DFL损失增益", value=1.5)
            pose_input = gr.Number(label="姿态损失增益", value=12.0)
            kobj_input = gr.Number(label="关键点目标损失增益", value=2.0)
            label_smoothing_input = gr.Number(label="标签平滑", value=0.0)
            nbs_input = gr.Number(label="标称批量大小", value=64)
            overlap_mask_input = gr.Checkbox(label="重叠遮罩", value=True)
            mask_ratio_input = gr.Number(label="遮罩比例", value=4)
            dropout_input = gr.Number(label="Dropout率", value=0.0)
            
            # 新增的高级选项
            augment_input = gr.Checkbox(label="使用数据增强", value=True)
            mixup_input = gr.Slider(minimum=0, maximum=1, value=0.1, label="Mixup 强度")
            copy_paste_input = gr.Slider(minimum=0, maximum=1, value=0.1, label="Copy-Paste 强度")
            mosaic_input = gr.Slider(minimum=0, maximum=1, value=1.0, label="Mosaic 强度")
            
            # 新增的 AMP 选项
            amp_input = gr.Checkbox(label="使用自动混合精度(AMP)", value=True)
            
        train_button = gr.Button("开始训练")
        train_output = gr.Textbox(label="训练结果")
        
        train_task_input.change(update_model_choices, inputs=[train_task_input], outputs=[train_model_input])
        train_button.click(train_model, 
                           inputs=[data_input, epochs_input, train_task_input, train_model_input, 
                                   device_input, include_test, test_size, val_size, batch_size_input, imgsz_input,
                                   patience_input, optimizer_input, lr0_input, lrf_input, momentum_input,
                                   weight_decay_input, warmup_epochs_input, warmup_momentum_input,
                                   warmup_bias_lr_input, box_input, cls_input, dfl_input, pose_input,
                                   kobj_input, label_smoothing_input, nbs_input, overlap_mask_input,
                                   mask_ratio_input, dropout_input, 
                                   # 新增的输入
                                   augment_input, mixup_input, copy_paste_input, mosaic_input, amp_input], 
                           outputs=train_output)
    
    with gr.Tab("验证"):
        val_task_input = gr.Dropdown(["检测", "分割", "姿态", "分类", "OBB"], label="任务类型")
        val_model_input = gr.Dropdown([], label="选择模型")
        val_data_input = gr.Textbox(label="验证数据集路径(yaml文件路径)")
        val_batch_size = gr.Number(label="批量大小", value=32)
        val_imgsz = gr.Number(label="图像尺寸", value=640)
        
        with gr.Accordion("高级选项", open=False):
            val_conf = gr.Slider(minimum=0.1, maximum=1.0, value=0.25, label="置信度阈值")
            val_iou = gr.Slider(minimum=0.1, maximum=1.0, value=0.7, label="IOU阈值")
            val_max_det = gr.Number(label="最大检测数", value=300)
            val_half = gr.Checkbox(label="使用半精度", value=False)
            val_plots = gr.Checkbox(label="生成图表", value=True)

        val_button = gr.Button("开始验证")
        val_output = gr.Textbox(label="验证结果")
        
        val_task_input.change(update_model_choices, inputs=[val_task_input], outputs=[val_model_input])
        val_button.click(validate_model, 
                         inputs=[val_data_input, val_task_input, val_model_input, device_input, 
                                 val_batch_size, val_imgsz, val_conf, val_iou, val_max_det, val_half, val_plots], 
                         outputs=val_output)
    
    with gr.Tab("预测"):
        with gr.Row():
            with gr.Column(scale=1):
                pred_task_input = gr.Dropdown(["检测", "分割", "姿态", "分类", "OBB"], label="任务类型")
                pred_model_input = gr.Dropdown([], label="选择模型")
                image_input = gr.Image(type="numpy", label="输入图像", scale=1, height=400, width="100%")
                
                with gr.Accordion("高级选项", open=False):
                    pred_conf = gr.Slider(minimum=0.1, maximum=1.0, value=0.25, label="置信度阈值")
                    pred_iou = gr.Slider(minimum=0.1, maximum=1.0, value=0.7, label="IOU阈值")
                    pred_max_det = gr.Number(label="最大检测数", value=300)
                    pred_classes = gr.Textbox(label="类别过滤（逗号分隔的索引）", value="")
                    pred_agnostic_nms = gr.Checkbox(label="类别无关NMS", value=False)
                    pred_augment = gr.Checkbox(label="使用TTA", value=False)

                predict_button = gr.Button("开始预测")

            with gr.Column(scale=1):
                image_output = gr.Image(type="numpy", label="预测结果", scale=1, height=400, width="100%")
                text_output = gr.JSON(label="预测详情")
                pose_data_file = gr.File(label="📊 姿态关键点数据 (Excel)", visible=False)
                pred_error_output = gr.Textbox(label="错误信息", visible=False)

        pred_task_input.change(update_model_choices, inputs=[pred_task_input], outputs=[pred_model_input])
        
        def process_prediction(image, task, model_name, device, conf, iou, max_det, classes, agnostic_nms, augment):
            plot_result, json_result, excel_path = predict_image(image, task, model_name, device, conf, iou, max_det, classes, agnostic_nms, augment)
            if plot_result is None and isinstance(json_result, dict) and "error" in json_result:
                return gr.update(visible=True, value=json_result["error"]), gr.update(visible=False), gr.update(visible=False), gr.update(visible=False, value=None)
            else:
                # 姿态任务时显示数据文件下载
                if excel_path:
                    return gr.update(visible=False), plot_result, json_result, gr.update(visible=True, value=excel_path)
                else:
                    return gr.update(visible=False), plot_result, json_result, gr.update(visible=False, value=None)

        predict_button.click(process_prediction, 
                            inputs=[image_input, pred_task_input, pred_model_input, device_input,
                                    pred_conf, pred_iou, pred_max_det, pred_classes, pred_agnostic_nms, pred_augment], 
                            outputs=[pred_error_output, image_output, text_output, pose_data_file])
    
    with gr.Tab("导出"):
        export_task_input = gr.Dropdown(["检测", "分割", "姿态", "分类", "OBB"], label="任务类型")
        export_model_input = gr.Dropdown([], label="选择模型")
        format_input = gr.Dropdown(["onnx", "openvino", "engine", "coreml", "saved_model", "pb", "tflite", "edgetpu", "tfjs"], label="导出格式")
        
        with gr.Accordion("高级选项", open=False):
            export_imgsz = gr.Number(label="图像尺寸", value=640)
            export_batch = gr.Number(label="批量大小", value=1)
            export_half = gr.Checkbox(label="使用半精度", value=False)
            export_int8 = gr.Checkbox(label="INT8量化", value=False)
            export_dynamic = gr.Checkbox(label="动态轴", value=False)
            export_simplify = gr.Checkbox(label="简化模型", value=True)
            export_opset = gr.Number(label="ONNX操作集", value=17)

        export_button = gr.Button("导出模型")
        export_output = gr.Textbox(label="导出结果")
        
        export_task_input.change(update_model_choices, inputs=[export_task_input], outputs=[export_model_input])
        export_button.click(export_model, 
                            inputs=[format_input, export_task_input, export_model_input, device_input,
                                    export_imgsz, export_batch, export_half, export_int8, export_dynamic, export_simplify, export_opset], 
                            outputs=export_output)

    with gr.Tab("视频推理"):
        with gr.Row():
            with gr.Column(scale=1):
                # 视频推理不支持分类任务，移除"分类"选项
                video_task_input = gr.Dropdown(["检测", "分割", "姿态", "OBB"], label="任务类型")
                video_model_input = gr.Dropdown([], label="选择模型")
                video_input = gr.Video(label="输入视频", scale=1, height=400, width="100%")
                
                with gr.Accordion("高级选项", open=False):
                    video_conf = gr.Slider(minimum=0.1, maximum=1.0, value=0.25, label="置信度阈值")
                    video_iou = gr.Slider(minimum=0.1, maximum=1.0, value=0.7, label="IOU阈值")
                    video_max_det = gr.Number(label="最大检测数", value=300)
                    video_classes = gr.Textbox(label="类别过滤（逗号分隔的索引）", value="")
                    video_agnostic_nms = gr.Checkbox(label="类别无关NMS", value=False)
                    video_augment = gr.Checkbox(label="使用TTA", value=False)

                video_button = gr.Button("开始视频推理")

            with gr.Column(scale=1):
                video_output = gr.Video(label="推理结果", scale=1, height=400, width="100%")
                video_pose_data_file = gr.File(label="📊 姿态关键点数据 (Excel)", visible=False)
                video_error_output = gr.Textbox(label="错误信息")

        video_task_input.change(update_model_choices, inputs=[video_task_input], outputs=[video_model_input])
        video_button.click(on_video_button_click, 
                           inputs=[video_input, video_task_input, video_model_input, device_input,
                                   video_conf, video_iou, video_max_det, video_classes, video_agnostic_nms, video_augment], 
                           outputs=[video_output, video_error_output, video_pose_data_file])

    with gr.Tab("数据导出"):
        gr.Markdown("""
        ### 📊 姿态估计数据导出
        从 YOLOv11 姿态估计模型的推理结果中导出完整原始数据到 Excel 文件。
        
        **导出内容包括：**
        - 边界框 (Bounding Box)：左上角和右下角像素坐标 [x1, y1, x2, y2]
        - 检测置信度 (Detection Confidence)
        - 17个 COCO 关键点：像素坐标 x, y + 关键点置信度 + 标准名称
        
        **支持输入：** 单张图像 / 视频文件（逐帧提取）
        
        > 💡 提示：在「预测」和「视频推理」Tab 中使用姿态任务时，也会自动生成 Excel 数据文件。
        """)
        
        with gr.Row():
            with gr.Column(scale=1):
                export_data_model_input = gr.Dropdown(
                    choices=get_available_models("姿态"),
                    label="选择姿态估计模型",
                    value=get_available_models("姿态")[0] if get_available_models("姿态") else None
                )
                
                with gr.Tabs("输入源"):
                    with gr.Tab("图像"):
                        export_data_image_input = gr.Image(type="numpy", label="上传图像", height=300)
                    with gr.Tab("视频"):
                        export_data_video_input = gr.Video(label="上传视频", height=300)
                
                export_data_source_type = gr.Radio(["image", "video"], label="输入源类型", value="image",
                                                    info="选择上传的是图像还是视频")
                
                with gr.Accordion("推理参数", open=True):
                    export_data_conf = gr.Slider(minimum=0.1, maximum=1.0, value=0.25, step=0.05, label="置信度阈值")
                    export_data_iou = gr.Slider(minimum=0.1, maximum=1.0, value=0.7, step=0.05, label="IOU阈值")
                    export_data_max_det = gr.Number(label="最大检测数", value=300)
                
                with gr.Accordion("导出设置", open=True):
                    export_data_output_dir = gr.Textbox(
                        label="导出目录路径",
                        value=os.path.join(os.getcwd(), "exports"),
                        info="留空则默认导出到项目目录下的 exports 文件夹"
                    )
                
                export_data_button = gr.Button("🚀 开始推理并导出数据", variant="primary")

            with gr.Column(scale=1):
                export_data_result = gr.Textbox(label="导出结果", lines=12, interactive=False)
                export_data_file = gr.File(label="下载导出文件")

        # 根据输入源类型切换显示
        export_data_source_type.change(
            fn=lambda t: (gr.update(visible=(t == "image")), gr.update(visible=(t == "video"))),
            inputs=[export_data_source_type],
            outputs=[export_data_image_input, export_data_video_input]
        )

        def _handle_export_data(image, video, source_type, model_name, device, conf, iou, max_det, output_dir):
            """处理数据导出按钮点击事件"""
            if source_type == "image":
                if image is None:
                    return "❌ 请先上传图像文件。", None
                source = image
            elif source_type == "video":
                if video is None:
                    return "❌ 请先上传视频文件。", None
                source = video
            else:
                return "❌ 未知的输入源类型。", None

            if not model_name:
                return "❌ 请先选择模型。", None

            excel_path, summary = export_pose_data(
                source=source,
                source_type=source_type,
                model_name=model_name,
                device=device,
                conf=conf,
                iou=iou,
                max_det=int(max_det),
                output_dir=output_dir
            )

            if excel_path is None:
                return summary, None
            return summary, excel_path

        export_data_button.click(
            _handle_export_data,
            inputs=[
                export_data_image_input, export_data_video_input, export_data_source_type,
                export_data_model_input, device_input,
                export_data_conf, export_data_iou, export_data_max_det,
                export_data_output_dir
            ],
            outputs=[export_data_result, export_data_file]
        )

if __name__ == "__main__":
    demo.launch(
        server_name="127.0.0.1", 
        server_port=7860, 
        inbrowser=True,
        share=False,  # 禁用 Gradio 共享链接
    )

# 如果您想完全禁用所有网络请求,可以在文件顶部添加以下代码
# os.environ['GRADIO_ANALYTICS_ENABLED'] = 'False'
