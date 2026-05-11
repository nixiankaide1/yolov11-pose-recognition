"""Multi-sheet biomechanics Excel export with formatted headers."""

import os
import logging
import numpy as np
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.config import COCO_KEYPOINT_NAMES, COCO_KEYPOINT_NAMES_CN, EXPORTS_DIR


# Color palette for headers
COLOR_FRAME = "E8D5F5"
COLOR_DETECT = "D9E1F2"
COLOR_BBOX = "E2EFDA"
COLOR_CONF = "FFF2CC"
COLOR_UPPER = "DDEBF7"   # blue
COLOR_LOWER = "E2EFDA"   # green
COLOR_TRUNK = "FCE4D6"   # orange
COLOR_COM = "FFF2CC"     # yellow
COLOR_KINEMATICS = "D9E1F2"  # blue-gray
COLOR_SUMMARY = "F4B4C2"  # pink/red
BORDER_COLOR = "BFBFBF"

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
header_font = Font(size=10, bold=True)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_align = Alignment(horizontal="center", vertical="center")


def _make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _style_data_cells(ws, start_row, total_cols):
    """Apply border and alignment to data rows."""
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_align
            cell.border = thin_border
            if isinstance(cell.value, float):
                cell.number_format = '0.0000'


def _auto_width(ws, total_cols, first_col_width=28):
    """Set column widths."""
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = first_col_width if col_idx == 1 else 12


def write_keypoints_sheet(ws, all_rows):
    """
    Sheet: 关键点数据 — same format as the original _write_pose_excel.
    all_rows: list of lists, each row = [image_name, frame_num, time_sec, person_id,
               x1, y1, x2, y2, det_conf, kp0_x, kp0_y, kp0_conf, ...]
    """
    ws.title = "关键点数据"
    base_col_count = 9
    kp_col_count = len(COCO_KEYPOINT_NAMES) * 3
    total_cols = base_col_count + kp_col_count

    # Row 1: merged category headers
    row1_groups = [
        (1, 1, "图像名称", COLOR_FRAME),
        (2, 3, "帧信息", COLOR_FRAME),
        (4, 4, "检测信息", COLOR_DETECT),
        (5, 8, "边界框", COLOR_BBOX),
        (9, 9, "检测置信度", COLOR_CONF),
    ]
    for kp_idx, kp_name in enumerate(COCO_KEYPOINT_NAMES):
        start_col = base_col_count + 1 + kp_idx * 3
        end_col = start_col + 2
        cn_name = COCO_KEYPOINT_NAMES_CN.get(kp_name, kp_name)
        label = f"{kp_name}\n({cn_name})"
        color = "FCE4D6" if kp_idx % 2 == 0 else "DDEBF7"
        row1_groups.append((start_col, end_col, label, color))

    for sc, ec, title, color in row1_groups:
        if sc != ec:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        cell = ws.cell(row=1, column=sc)
        cell.value = title
        cell.font = header_font
        cell.fill = _make_fill(color)
        cell.alignment = header_align
        cell.border = thin_border
        for c in range(sc, ec + 1):
            wc = ws.cell(row=1, column=c)
            wc.fill = _make_fill(color)
            wc.border = thin_border

    # Row 2: sub-headers
    sub_headers = ["图像名称", "帧编号", "时间 (s)", "检测ID", "x1 (px)", "y1 (px)", "x2 (px)", "y2 (px)", "检测置信度"]
    for kp_name in COCO_KEYPOINT_NAMES:
        sub_headers.extend(["x (px)", "y (px)", "conf"])
    sub_colors = [COLOR_FRAME]*4 + [COLOR_BBOX]*4 + [COLOR_CONF]
    for kp_idx in range(len(COCO_KEYPOINT_NAMES)):
        color = "FCE4D6" if kp_idx % 2 == 0 else "DDEBF7"
        sub_colors.extend([color, color, color])

    for col_idx, (h, color) in enumerate(zip(sub_headers, sub_colors), start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = _make_fill(color)
        cell.alignment = header_align
        cell.border = thin_border

    for row_data in all_rows:
        ws.append(row_data)

    _style_data_cells(ws, 3, total_cols)
    _auto_width(ws, total_cols)
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20


def write_angles_sheet(ws, angle_rows):
    """
    Sheet: 关节角度
    angle_rows: list of lists, each row =
        [image_name, frame_num, time_sec, person_id,
         left_elbow, right_elbow, left_shoulder, right_shoulder,
         left_hip, right_hip, left_knee, right_knee,
         left_ankle, right_ankle, trunk_inclination, neck_angle]
    """
    ws.title = "关节角度"
    angle_names = [
        "left_elbow", "right_elbow", "left_shoulder", "right_shoulder",
        "left_hip", "right_hip", "left_knee", "right_knee",
        "left_ankle", "right_ankle", "trunk_inclination", "neck_angle",
    ]
    cn_names = ["左肘", "右肘", "左肩", "右肩", "左髋", "右髋", "左膝", "右膝", "左踝", "右踝", "躯干倾角", "颈部角度"]

    total_cols = 4 + len(angle_names)

    # Row 1: category headers
    row1_groups = [
        (1, 1, "图像名称", COLOR_FRAME),
        (2, 3, "帧信息", COLOR_FRAME),
        (4, 4, "人体ID", COLOR_DETECT),
        (5, 8, "上肢角度", COLOR_UPPER),
        (9, 14, "下肢角度", COLOR_LOWER),
        (15, 16, "躯干/颈部", COLOR_TRUNK),
    ]
    for sc, ec, title, color in row1_groups:
        if sc != ec:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        cell = ws.cell(row=1, column=sc)
        cell.value = title
        cell.font = header_font
        cell.fill = _make_fill(color)
        cell.alignment = header_align
        cell.border = thin_border
        for c in range(sc, ec + 1):
            wc = ws.cell(row=1, column=c)
            wc.fill = _make_fill(color)
            wc.border = thin_border

    # Row 2: sub-headers
    sub_headers = ["图像名称", "帧编号", "时间 (s)", "人体ID"]
    sub_colors = [COLOR_FRAME, COLOR_FRAME, COLOR_FRAME, COLOR_DETECT]
    for i, an in enumerate(angle_names):
        sub_headers.append(f"{cn_names[i]} (°)")
        if i < 4:
            sub_colors.append(COLOR_UPPER)
        elif i < 10:
            sub_colors.append(COLOR_LOWER)
        else:
            sub_colors.append(COLOR_TRUNK)

    for col_idx, (h, color) in enumerate(zip(sub_headers, sub_colors), start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = _make_fill(color)
        cell.alignment = header_align
        cell.border = thin_border

    for row_data in angle_rows:
        ws.append(row_data)

    _style_data_cells(ws, 3, total_cols)
    _auto_width(ws, total_cols, 24)
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20


def write_com_sheet(ws, com_rows):
    """
    Sheet: 质心数据
    com_rows: list of lists, each row =
        [image_name, frame_num, time_sec, com_x, com_y, com_vx, com_vy, com_speed, com_ax, com_ay]
    """
    ws.title = "质心数据"
    headers = ["图像名称", "帧编号", "时间 (s)", "COM_x (px)", "COM_y (px)",
               "COM_vx (px/s)", "COM_vy (px/s)", "COM速度 (px/s)", "COM_ax (px/s²)", "COM_ay (px/s²)"]
    total_cols = len(headers)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(size=10, bold=True)
        cell.fill = _make_fill(COLOR_COM)
        cell.alignment = header_align
        cell.border = thin_border

    for row_data in com_rows:
        ws.append(row_data)

    _style_data_cells(ws, 2, total_cols)
    _auto_width(ws, total_cols, 24)
    ws.freeze_panes = "A2"


def write_kinematics_sheet(ws, kinematics_rows):
    """
    Sheet: 运动学数据 (long format)
    kinematics_rows: list of lists, each row =
        [image_name, frame_num, time_sec, joint_name, joint_cn_name, angle, angular_vel, angular_acc]
    """
    ws.title = "运动学数据"
    headers = ["图像名称", "帧编号", "时间 (s)", "关节名称", "关节中文名", "角度 (°)", "角速度 (°/s)", "角加速度 (°/s²)"]
    total_cols = len(headers)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(size=10, bold=True)
        cell.fill = _make_fill(COLOR_KINEMATICS)
        cell.alignment = header_align
        cell.border = thin_border

    for row_data in kinematics_rows:
        ws.append(row_data)

    _style_data_cells(ws, 2, total_cols)
    _auto_width(ws, total_cols, 24)
    ws.freeze_panes = "A2"


def write_statistics_sheet(ws, statistics_rows):
    """
    Sheet: 统计摘要
    statistics_rows: list of dicts
    """
    ws.title = "统计摘要"
    headers = ["关节名称", "平均值 (°)", "最大值 (°)", "最小值 (°)", "活动范围 ROM (°)", "标准差 (°)", "有效帧数"]
    total_cols = len(headers)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(size=10, bold=True, color="FFFFFF")
        cell.fill = _make_fill("C0392B")
        cell.alignment = header_align
        cell.border = thin_border

    for row_data in statistics_rows:
        row_list = [row_data.get(h, "N/A") for h in headers]
        ws.append(row_list)

    _style_data_cells(ws, 2, total_cols)
    _auto_width(ws, total_cols, 18)
    ws.freeze_panes = "A2"


def write_reference_sheet(ws):
    """Sheet: 关键点说明"""
    ws.title = "关键点说明"
    ws.append(["关键点编号", "英文名称", "中文名称", "分组"])
    group_names = {
        "nose": "头部", "left_eye": "头部", "right_eye": "头部",
        "left_ear": "头部", "right_ear": "头部",
        "left_shoulder": "上肢", "right_shoulder": "上肢",
        "left_elbow": "上肢", "right_elbow": "上肢",
        "left_wrist": "上肢", "right_wrist": "上肢",
        "left_hip": "下肢", "right_hip": "下肢",
        "left_knee": "下肢", "right_knee": "下肢",
        "left_ankle": "下肢", "right_ankle": "下肢",
    }
    for idx, name in enumerate(COCO_KEYPOINT_NAMES):
        cn = COCO_KEYPOINT_NAMES_CN.get(name, name)
        group = group_names.get(name, "")
        ws.append([idx, name, cn, group])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10


def write_biomechanics_excel(keypoint_rows, angle_rows, com_rows, kinematics_rows, statistics_rows, prefix="analysis"):
    """
    Create a multi-sheet biomechanics Excel workbook.

    Returns the output file path.
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active

    if keypoint_rows:
        write_keypoints_sheet(ws1, keypoint_rows)
    else:
        ws1.title = "关键点数据"
        ws1.append(["无关键点数据"])

    if angle_rows:
        ws2 = wb.create_sheet()
        write_angles_sheet(ws2, angle_rows)

    if com_rows:
        ws3 = wb.create_sheet()
        write_com_sheet(ws3, com_rows)

    if kinematics_rows:
        ws4 = wb.create_sheet()
        write_kinematics_sheet(ws4, kinematics_rows)

    if statistics_rows:
        ws5 = wb.create_sheet()
        write_statistics_sheet(ws5, statistics_rows)

    ws_ref = wb.create_sheet()
    write_reference_sheet(ws_ref)

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORTS_DIR, f"biomechanics_{prefix}_{timestamp}.xlsx")
    wb.save(filepath)
    logging.info(f"生物力学Excel已导出: {filepath}")
    return filepath
